#-----------------------------------------------------------
#ライブリの読込
#-----------------------------------------------------------
import openpyxl #エクセルのブックを操作するライブラリの読み込み
import re #文字列操作をするライブラリの読み込み

#-----------------------------------------------------------
#コピペする関数定義 copyAndPaste
# 第一引数　bookName：対象のExcelブック名
# 第二引数　sheetName：シート名
# 第三引数　fromCellRange：コピー元セル範囲
# 第四引数　toCell：コピー先セル番地
#-----------------------------------------------------------
def copyAndPaste(bookName, sheetName, fromCellRange, toCell):
    #指定したファイル名でエクセルブックを開く
    workBook = openpyxl.load_workbook(bookName)

    #指定したシート名のシートを取得する
    workSheet = workBook[sheetName]

    fromStartCell = "" #コピー元開始セル
    fromEndCell = "" #コピー元終了セル

    if fromCellRange.find(':') != -1:
        #:が含まれているかどうか判定
        workCellData = re.split(":", fromCellRange)
        fromStartCell = workCellData[0]
        fromEndCell = workCellData[1]

        #セル番地から行番号と列番号を取得
        fromStartCell = getRowAndColumn(fromStartCell)
        fromEndCell = getRowAndColumn(fromEndCell)
    else:
        #:が含まれていないときに処理される
        fromStartCell = getRowAndColumn(fromCellRange)
        fromEndCell = fromStartCell

    toCell = getRowAndColumn(toCell)

    #行ループ
    for i in range(fromStartCell[0],fromEndCell[0] + 1):
        #列ループ
        for j in range(fromStartCell[1], fromEndCell[1] + 1):
            workSheet.cell(row = toCell[0] - 1 + i, column = toCell[1] - 1 + j).value = workSheet.cell(row = i, column = j).value

    #ブックを保存する
    workBook.save(bookName)

#-----------------------------------------------------------
#指定したセル名から行番号と列番号を取得する関数定義 getRowAndColumn
# 第一引数　
#-----------------------------------------------------------
def getRowAndColumn(cellAdress):

    #列名の取り出し（リスト）
    column = re.split("[\d]", cellAdress)
    #列名を取り出し（文字列）
    column = column[0]
    #行番号を取得
    row = cellAdress.replace(column, "")
    #行番号を文字列から数値に変換
    row = int(row)
    #列名を列番号に変換
    column = openpyxl.utils.column_index_from_string(column)

    #行番号と列番号をリストで戻す
    return [row, column]

#-----------------------------------------------------------
#メイン処理
#-----------------------------------------------------------
#関数の呼び出し　A1からD4の範囲をA10に貼り付け
copyAndPaste("sample.xlsx","Sheet1","A1:D4","A10")
